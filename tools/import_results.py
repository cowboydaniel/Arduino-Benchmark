#!/usr/bin/env python3
"""
import_results.py — Import Arduino benchmark CSV output into the results spreadsheet.

Usage:
    python tools/import_results.py <csv_file> <xlsx_file> [--dry-run]

Arguments:
    csv_file    Path to the CSV output saved from Serial Monitor (CSV_OUTPUT=1 mode).
    xlsx_file   Path to the results spreadsheet to update.
                If omitted and the CSV contains "Part,1" or "Part,2", the matching
                spreadsheet is auto-detected in the current directory.
    --dry-run   Print what would be written without modifying the spreadsheet.

The CSV file must be in the format produced by the benchmark sketch with CSV_OUTPUT=1:
    Part,1
    Board,Arduino Uno
    Benchmark Version,1.0.0
    CPU (MHz),16
    Addition (ops/ms),399.74
    ...

The script will:
  1. Find the column for the board (by name in row 1), or create a new column.
  2. For each label in the CSV, find the matching row in column A of the spreadsheet.
  3. Write the numeric value into the board's cell.
  4. Warn about labels present in the CSV but missing from the spreadsheet.
  5. Leave spreadsheet cells untouched for blank CSV values (N/A metrics).
"""

import sys
import os
import re
import argparse
import glob as globmod


def parse_csv(path):
    """Parse the benchmark CSV file into an ordered list of (label, value) pairs.

    Returns a dict with 'rows' (list of (label, value_str)) and 'meta'
    (dict of well-known header fields: Part, Board, Benchmark Version, etc.)
    Blank-value rows (label present, value empty) are included as (label, None).
    """
    rows = []
    meta = {}
    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            # Split on first comma only
            parts = line.split(",", 1)
            label = parts[0].strip()
            value = parts[1].strip() if len(parts) > 1 else ""
            if not label:
                continue
            # Track well-known metadata rows
            if label in ("Part", "Board", "Benchmark Version"):
                meta[label] = value
            rows.append((label, value if value else None))
    return {"rows": rows, "meta": meta}


def try_numeric(value_str):
    """Convert a string to int or float if possible, otherwise return the string."""
    if value_str is None:
        return None
    # Try integer first (no decimal point)
    try:
        v = int(value_str)
        return v
    except ValueError:
        pass
    # Try float
    try:
        v = float(value_str)
        return v
    except ValueError:
        pass
    return value_str


def find_spreadsheet(part, search_dir="."):
    """Auto-detect the spreadsheet for the given part number (1 or 2)."""
    prefix = f"Part{part}_"
    pattern = os.path.join(search_dir, f"{prefix}*_Results_v*.xlsx")
    matches = globmod.glob(pattern)
    if not matches:
        return None
    # Prefer highest version number (simple string sort works for semver x.y.z)
    matches.sort(reverse=True)
    return matches[0]


def import_results(csv_path, xlsx_path, dry_run=False):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
        sys.exit(1)

    # ── Parse CSV ──────────────────────────────────────────────────────────────
    data = parse_csv(csv_path)
    rows = data["rows"]
    meta = data["meta"]

    board_name = meta.get("Board")
    bench_version = meta.get("Benchmark Version")
    part = meta.get("Part")

    if not board_name:
        print("ERROR: CSV does not contain a 'Board' row. Is this a CSV_OUTPUT=1 file?")
        sys.exit(1)

    print(f"Board:             {board_name}")
    print(f"Benchmark Version: {bench_version or '(not found)'}")
    print(f"Part:              {part or '(not found)'}")
    print(f"Metrics in CSV:    {sum(1 for _, v in rows if v is not None)} values, "
          f"{sum(1 for _, v in rows if v is None)} blanks")
    print()

    # ── Auto-detect spreadsheet ────────────────────────────────────────────────
    if not xlsx_path:
        if not part:
            print("ERROR: CSV has no 'Part' row and no xlsx path was given.")
            sys.exit(1)
        search_dir = os.path.dirname(os.path.abspath(csv_path)) or "."
        # Also try the repo root (one level up from tools/)
        candidates = [search_dir, os.path.join(search_dir, ".."), "."]
        xlsx_path = None
        for d in candidates:
            found = find_spreadsheet(part, d)
            if found:
                xlsx_path = os.path.abspath(found)
                break
        if not xlsx_path:
            print(f"ERROR: Could not find a Part {part} spreadsheet. "
                  f"Pass the xlsx path explicitly.")
            sys.exit(1)
        print(f"Auto-detected spreadsheet: {xlsx_path}")

    if not os.path.exists(xlsx_path):
        print(f"ERROR: Spreadsheet not found: {xlsx_path}")
        sys.exit(1)

    # ── Load workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    print(f"Opened: {xlsx_path}  (sheet: '{ws.title}', "
          f"{ws.max_row} rows × {ws.max_column} cols)")
    print()

    # ── Build row-label → row-number lookup from column A ────────────────────
    label_to_row = {}
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None:
            label_to_row[str(cell_val).strip()] = r

    # ── Find or create the board column ───────────────────────────────────────
    board_col = None
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if header is not None and str(header).strip() == board_name:
            board_col = c
            break

    if board_col is None:
        board_col = ws.max_column + 1
        print(f"Board '{board_name}' not found — creating new column {board_col}.")
        if not dry_run:
            ws.cell(row=1, column=board_col).value = board_name
    else:
        print(f"Board '{board_name}' found in column {board_col}.")

    # Write benchmark version into the "Version" row if it exists
    if bench_version and "Version" in label_to_row:
        ver_row = label_to_row["Version"]
        if not dry_run:
            ws.cell(row=ver_row, column=board_col).value = bench_version
        print(f"  Set Version row ({ver_row}) = {bench_version}")

    print()

    # ── Write metric values ────────────────────────────────────────────────────
    written = 0
    skipped_blank = 0
    not_found = []

    # Skip pure metadata rows that aren't spreadsheet data rows
    meta_labels = {"Part", "Board", "Benchmark Version"}

    for label, value_str in rows:
        if label in meta_labels:
            continue
        if value_str is None:
            skipped_blank += 1
            continue  # N/A metric — leave cell untouched

        if label not in label_to_row:
            not_found.append((label, value_str))
            continue

        target_row = label_to_row[label]
        numeric_value = try_numeric(value_str)

        if dry_run:
            current = ws.cell(row=target_row, column=board_col).value
            print(f"  [DRY RUN] Row {target_row:4d}  {label!r:40s} "
                  f"{str(current or '(empty)'):>12} → {numeric_value}")
        else:
            ws.cell(row=target_row, column=board_col).value = numeric_value

        written += 1

    # ── Summary ────────────────────────────────────────────────────────────────
    print()
    print(f"Results: {written} values written, {skipped_blank} blanks skipped.")

    if not_found:
        print(f"\nWARNING: {len(not_found)} CSV label(s) not found in spreadsheet column A:")
        for lbl, val in not_found:
            print(f"  '{lbl}' = {val}")
        print("  These rows may need to be added to the spreadsheet manually.")

    # ── Save ───────────────────────────────────────────────────────────────────
    if not dry_run:
        wb.save(xlsx_path)
        print(f"\nSaved: {xlsx_path}")
    else:
        print("\n[DRY RUN] No changes saved.")


def main():
    parser = argparse.ArgumentParser(
        description="Import benchmark CSV output into the results spreadsheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("csv_file", help="CSV file from Serial Monitor (CSV_OUTPUT=1)")
    parser.add_argument(
        "xlsx_file",
        nargs="?",
        default=None,
        help="Results spreadsheet to update (auto-detected if omitted)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without modifying the spreadsheet",
    )
    args = parser.parse_args()

    import_results(args.csv_file, args.xlsx_file, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
